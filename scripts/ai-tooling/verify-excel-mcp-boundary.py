#!/usr/bin/env python3
"""Verify the Excel MCP filesystem boundary actually holds.

Two independent checks, because the boundary has two independent failure
modes and the README documents neither of them accurately:

  1. CONFINEMENT — get_excel_path() must reject absolute paths and any
     relative path that escapes EXCEL_FILES_PATH, including via symlink.
     This is the check that makes the difference between "an Excel tool" and
     "an arbitrary file read/write tool".

  2. LIVE READ/WRITE — the server must still actually work inside the
     boundary, on synthetic data only. A boundary that breaks the feature is
     not a win.

Run:  .venv-mcp/bin/python scripts/ai-tooling/verify-excel-mcp-boundary.py
"""
import os
import sys
import tempfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
WORKSPACE = REPO_ROOT / "data" / "mcp" / "excel"

failures = []
passes = []


def check(name, condition, detail=""):
    (passes if condition else failures).append(f"{name}{(' — ' + detail) if detail else ''}")
    print(f"  {'PASS' if condition else 'FAIL'}  {name}" + (f" — {detail}" if detail else ""))


def main():
    WORKSPACE.mkdir(parents=True, exist_ok=True)
    os.environ["EXCEL_FILES_PATH"] = str(WORKSPACE)

    try:
        import excel_mcp.server as srv
    except ImportError:
        print("excel-mcp-server is not installed in this interpreter.")
        print("Run: bash scripts/ai-tooling/start-excel-mcp.sh --check")
        return 2

    # The server only assigns EXCEL_FILES_PATH inside run_sse/run_streamable_http.
    # Assign it directly so we test the same code path those transports use.
    srv.EXCEL_FILES_PATH = str(WORKSPACE)

    print("\n1. CONFINEMENT (EXCEL_FILES_PATH = data/mcp/excel)")

    def rejected(candidate):
        try:
            srv.get_excel_path(candidate)
            return False
        except ValueError:
            return True

    check("absolute path /etc/passwd rejected", rejected("/etc/passwd"))
    check("absolute path to repo .env rejected", rejected(str(REPO_ROOT / ".env")))
    check("traversal ../../.env rejected", rejected("../../.env"))
    check("traversal ../../../../etc/passwd rejected", rejected("../../../../etc/passwd"))
    check("deep traversal to db.sqlite3 rejected", rejected("../../db.sqlite3"))
    check("NUL byte in filename rejected", rejected("a\x00b.xlsx"))
    check("empty filename rejected", rejected(""))

    # Symlink escape: realpath() on both sides is what defeats this.
    link = WORKSPACE / "escape-link"
    try:
        if link.is_symlink() or link.exists():
            link.unlink()
        link.symlink_to(REPO_ROOT)
        check("symlink escape rejected", rejected("escape-link/.env"))
    finally:
        if link.is_symlink():
            link.unlink()

    # A legitimate in-workspace path must still resolve, or the boundary is
    # just breakage rather than confinement.
    try:
        resolved = srv.get_excel_path("reports/q1.xlsx")
        inside = Path(resolved).is_relative_to(WORKSPACE.resolve())
        check("legitimate relative path accepted and inside workspace", inside, resolved)
    except ValueError as exc:
        check("legitimate relative path accepted", False, str(exc))

    print("\n2. STDIO TRANSPORT (documented as unconfined — confirming)")
    srv.EXCEL_FILES_PATH = None
    try:
        leaked = srv.get_excel_path("/etc/passwd")
        check(
            "stdio mode confirmed UNCONFINED (this is why it is never used)",
            leaked == "/etc/passwd",
            f"returned {leaked}",
        )
    except ValueError:
        check("stdio mode unexpectedly confined", False, "upstream behaviour changed — re-audit")
    srv.EXCEL_FILES_PATH = str(WORKSPACE)

    print("\n3. LIVE READ/WRITE on synthetic data")
    from excel_mcp.workbook import create_workbook
    from excel_mcp.data import write_data
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory(dir=WORKSPACE) as tmp:
        book = Path(tmp) / "synthetic-emissions.xlsx"
        create_workbook(str(book))
        # create_workbook names the default sheet "Sheet1", not "Sheet".
        # Passing a name that does not exist makes write_data create a second
        # sheet and leave wb.active pointing at the empty original — so read
        # the sheet name back rather than assuming it.
        sheet_name = load_workbook(book).sheetnames[0]
        rows = [
            ["site", "scope_1_tco2e", "scope_2_tco2e", "reporting_year"],
            ["Synthetic Plant A", 1240.5, 880.25, 2025],
            ["Synthetic Plant B", 640.0, 410.75, 2025],
        ]
        write_data(str(book), sheet_name, rows, start_cell="A1")
        wb = load_workbook(book)
        ws = wb[sheet_name]
        check("workbook created on disk", book.exists())
        check("header round-trips", ws["A1"].value == "site", repr(ws["A1"].value))
        check("numeric value round-trips", ws["B2"].value == 1240.5, repr(ws["B2"].value))
        check("row count correct", ws.max_row == 3, f"max_row={ws.max_row}")
        check("file written inside workspace", Path(book).resolve().is_relative_to(WORKSPACE.resolve()))

    print(f"\n{len(passes)} passed, {len(failures)} failed")
    if failures:
        print("\nFAILURES:")
        for f in failures:
            print(f"  - {f}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
